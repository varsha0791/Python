import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")  #reads excel

product_list = inv_file["Sheet1"] #list

#1. products per supplier
products_per_supplier = {}  #empty dictionary

#print max row
print(product_list.max_row)

#reading each row
for product_row in range(2, product_list.max_row + 1):      #range is used becoz for loop iterates from 0,1,... indexes and we will iterate from 2nd row and since excel starts from 1st row (having heading) skipping that and starting from 2nd row
 #it will read till only last but 1 row since we skipped 1st row so we are adding last row.. +1
       supplier_name = product_list.cell(product_row, 4).value #.value gets the suppliers name , 4 becoz supplier name is in D column

       if supplier_name in products_per_supplier:      #checking if supplier name already exists in dictionary then go to next supplier i.e increments and go to next
          current_num_products = products_per_supplier[supplier_name]
          products_per_supplier[supplier_name] = current_num_products + 1
       else:
          print("adding a new supplier") #3 times printed for 3 suppliers as 3 iterations
          products_per_supplier[supplier_name] = 1 #for first time when dictionary is empty then it set product count as 1

print(products_per_supplier)